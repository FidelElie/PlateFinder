"""
For creating excel spreadsheets for experimental aspect
"""
import openpyxl as xl

class ExcelHandler(object):
    """Creates excel sheet object for parsing data to excel"""

    workbook = xl.Workbook()

    def __init__(self, data):
        self.data = data
        self.sheet = self.workbook.active
        self.standard_heads = ["Body", "Plate", "Calculated X", "Calculated Y", "Measured X", "Measured Y", "Delta X", "Delta Y", "D X - Average", "D Y - Average", "Av Diff ^ 2 (x)", "Av Diff ^ 2(Y)", "Comments" ]
        self.results_heads = ["Average Delta X", "Average Delta Y", "Standard Deviation X", "Standard Deviation Y", "Standard Error Delta X", "Standard Error Delta Y"]
        self.sheet.title = "Body Plate Results Sheet"

    def set_headings(self):
        """Sets headings of table"""
        head_fonts = xl.styles.Font(size=12, bold=True)
        for i in range(0, len(self.standard_heads)):
            self.sheet.cell(row=1, column=i+1,value = self.standard_heads[i])
            self.sheet.cell(row=1, column=i+1,value = self.standard_heads[i]).font = head_fonts

    def set_results(self, bottom_index):
        """Sets result headings"""
        result_fonts = xl.styles.Font(size=12, bold=True)
        starting_index = bottom_index + 3
        for i in range(0, len(self.results_heads)):
            self.sheet.cell(row= starting_index + i, column = 2, value = self.results_heads[i])
            self.sheet.cell(row= starting_index + i, column = 2, value = self.results_heads[i]).font = result_fonts
        return starting_index

    def set_initial_data(self):
        """Sets intial data to respective fields"""
        bottom_row = len(self.data) + 2
        for i in range(len(self.data)):
            for j in range(0, len(self.data[0])):
                self.sheet.cell(row=i + 2, column=j + 1, value=self.data[i][j])
        return bottom_row

    def set_formulas(self, starting_index):
        """Sets formulas to respective cells"""
        for i in range(0, len(self.data)):
            for j in range(5, len(self.standard_heads)):
                if j == 5:
                    self.sheet["G{}".format(i + 2)] = "=C{0}-E{0}".format(i + 2)
                elif j == 6:
                    self.sheet["H{}".format(i + 2)] = "=D{0}-F{0}".format(i + 2)
                elif j == 7:
                    self.sheet["I{}".format(i + 2)] = "=G{}-C{}".format(i + 2, starting_index)
                elif j == 8:
                    self.sheet["J{}".format(i + 2)] = "=H{}-C{}".format(i + 2, starting_index + 1)
                elif j == 9:
                    self.sheet["K{}".format(i + 2)] = "=I{}^2".format(i + 2)
                elif j == 10:
                    self.sheet["L{}".format(i + 2)] = "=J{}^2".format(i + 2)

        for i in range(len(self.results_heads)):
            if i == 0:
                self.sheet["C{}".format(starting_index)] = "=AVERAGE(G{}:G{})".format(2, len(self.data) + 1)
            elif i == 1:
                self.sheet["C{}".format(starting_index + 1)] = "=AVERAGE(H{}:H{})".format(2, len(self.data) + 1)
            elif i == 2:
                self.sheet["C{}".format(starting_index + 2)] = "=STDEV.P(K{}:K{})".format(2, len(self.data) + 1)
            elif i == 3:
                self.sheet["C{}".format(starting_index + 3)] = "=STDEV.P(L{}:L{})".format(2, len(self.data) + 1)
            elif i == 4:
                self.sheet["C{}".format(starting_index + 4)] = "=C{}/{}".format(starting_index + 2 ,len(self.data))
            elif i == 5:
                self.sheet["C{}".format(starting_index + 5)] =  "=C{}/{}".format(starting_index + 3 ,len(self.data))

    def scale_columns(self):
        """Scale the columns of the spreadsheet to fit"""
        for col in self.sheet.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.05
            self.sheet.column_dimensions[column].width = adjusted_width

    def parse_data(self):
        """Parses all data to the spreadsheet"""
        self.set_headings()
        bottom_index = self.set_initial_data()
        starting_index = self.set_results(bottom_index)
        self.set_formulas(starting_index)
        self.scale_columns()
        return self.workbook

if __name__ == '__main__':
    excel_handler = ExcelHandler([1])
    excel_handler.set_headings()
    test_workbook = excel_handler.return_excel_data()
    test_workbook.save('test.xlsx')
